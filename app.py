import streamlit as st
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.drawing.image import Image
from io import BytesIO
import time
import random

# Streamlit App
st.title("Amazon Product Scraper By Kumaran 🫶")
st.write("Provide a product link and the number of pages to scrape.")
st.write("Connect with me on [LinkedIn](https://www.linkedin.com/in/elangkumaran)")

# Input fields
product_link = st.text_input("Product Link", "https://www.amazon.in/s?k=mutton+pickle+250gm")
num_pages = st.number_input("Number of Pages to Scrape", min_value=1, max_value=10, value=4)

# Function to make requests with retries
def fetch_with_retries(url, headers, retries=5):
    for i in range(retries):
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()  # Raise an error for bad responses
            return response
        except requests.exceptions.RequestException as e:
            if i < retries - 1:  # Don't wait on the last attempt
                st.warning(f"Attempt {i + 1} failed: {e}. Retrying...")
                time.sleep(random.uniform(1, 3))  # Wait a random time between 1 to 3 seconds
            else:
                st.error(f"Error fetching the URL after {retries} attempts: {e}")
                return None
    return None

if st.button("Scrape"):
    # Create progress bar
    progress = st.progress(0)

    # Set up headers
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36",
        "Accept-Language": "en-US,en;q=0.9",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Referer": "https://www.amazon.in/",
    }

    # Initialize workbook
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Amazon Products"
    sheet.append(["Image", "Product Name", "Price", "Rating and Review", "Product Link"])  # Header row

    row = 2
    page_count = 0
    base_url = product_link

    while base_url and page_count < num_pages:
        # Update progress bar
        progress.progress((page_count + 1) / num_pages)
        st.write(f"Scraping page: {base_url}")

        response = fetch_with_retries(base_url, headers)
        if response is None:
            break  # Exit if the response is None after retries

        soup = BeautifulSoup(response.content, "html.parser")

        # Scrape product details
        products = soup.find_all("div", {"data-component-type": "s-search-result"})
        st.write(f"Found {len(products)} products on page {page_count + 1}")

        for product in products:
            # Extract image URL
            image_tag = product.find("img", class_="s-image")
            image_url = image_tag["src"] if image_tag else None

            # Extract product name
            name_tag = product.find("h2", {"class": "a-size-base-plus a-spacing-none a-color-base a-text-normal"})
            product_name = name_tag.text.strip() if name_tag else "N/A"

            # Extract price
            price_tag = product.find("span", class_="a-price-whole")
            product_price = price_tag.text.replace(",", "") if price_tag else "N/A"

            # Extract rating and review
            rating_tag = product.find("span", class_="a-icon-alt")
            review_count_tag = product.find("span", class_="a-size-base s-underline-text")
            product_rating = rating_tag.text if rating_tag else "N/A"
            review_count = review_count_tag.text if review_count_tag else "N/A"
            rating_and_review = f"{product_rating} ({review_count})"

            # Extract product link
            link_tag = name_tag.find_parent("a") if name_tag else None
            product_link = "https://www.amazon.in" + link_tag["href"] if link_tag else "N/A"

            # Embed the image into Excel
            if image_url:
                try:
                    img_data = requests.get(image_url).content
                    img = Image(BytesIO(img_data))
                    img.width = 80  # Resize to fit cell
                    img.height = 80
                    cell_address = f"A{row}"
                    sheet.add_image(img, cell_address)
                except Exception as e:
                    st.warning(f"Error downloading image: {e}")

            # Add other product details to the Excel sheet
            sheet.cell(row=row, column=2, value=product_name)  # Product name
            sheet.cell(row=row, column=3, value=product_price)  # Price
            sheet.cell(row=row, column=4, value=rating_and_review)  # Rating and review
            sheet.cell(row=row, column=5, value=product_link)  # Product link

            row += 1

        # Find the next page link
        next_page = soup.find("a", {"class": "s-pagination-next"})
        if next_page and "href" in next_page.attrs:
            base_url = "https://www.amazon.in" + next_page["href"]
            time.sleep(random.uniform(3, 5))  # Sleep for a random time between 3 to 5 seconds
        else:
            base_url = None
        
        page_count += 1

    # Save workbook to a BytesIO stream
    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    # Allow download of the Excel file
    st.download_button(
        label="Download Excel File",
        data=output,
        file_name="Amazon_Products.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    st.success("Scraping completed!")
